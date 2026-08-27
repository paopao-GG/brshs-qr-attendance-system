"""DepEd School Form 2 -- Daily Attendance Report of Learners.

The register in xlsx.py is SF2-*shaped*: students down the side, days across the top.
This is the form itself, laid out to the geometry of the LIS workbook the school
already submits (docs/SF2_2025_Grade-10-Year-IV-RESILIENT.xls) -- 47 columns, 25 day
slots, a male block above a female block, and the summary and signature panel.

Built rather than filled. The obvious alternative was to open the school's .xls as a
template and write into it, and it was rejected for one reason: that file has exactly
17 male and 22 female rows baked into its merges, and inserting rows for a class of 41
tears every merge and border below the insertion point. The geometry is a handful of
constants; the row count is not negotiable.

WHAT THIS FORM MEANS, because it is not the same arithmetic as the register:

  blank   present -- SF2 marks the exception, not the norm
  x       absent
  shaded  tardy (see LATE below)

Excused days are BLANK here and named in the Remarks column. That follows TRACKIFY's
own rule -- an excused day leaves the rate denominator rather than counting against a
student (docs/analytics-model.md section 1) -- so marking one 'x' would contradict the
register the same school gets from the same database on the same afternoon.

Nothing on this sheet is invented. Enrolment as of the first Friday, transfers and
drop-outs are not things TRACKIFY records, so their boxes are left blank for a person,
and the percentages that depend on them are left blank too.
"""

from __future__ import annotations

import sqlite3
from calendar import monthrange
from dataclasses import dataclass, field
from datetime import date as Date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

MONTHS = ("January", "February", "March", "April", "May", "June", "July",
          "August", "September", "October", "November", "December")

# --- geometry, read out of the school's own workbook -------------------------
# Every constant below was extracted from the BIFF records of
# docs/SF2_2025_Grade-10-Year-IV-RESILIENT.xls. Do not "tidy" them.

LAST_COLUMN = 47                         # AU

# The 25 day slots, as (first column, last column) 1-based inclusive pairs. The pairs
# are uneven because the LIS export splits some slots across two columns and not
# others; every slot still measures 2.86 characters, so they print identical.
DAY_SLOTS: tuple[tuple[int, int], ...] = (
    (6, 7), (8, 8), (9, 9), (10, 10), (11, 11),
    (12, 13), (14, 14), (15, 15), (16, 16), (17, 17),
    (18, 19), (20, 20), (21, 21), (22, 23), (24, 25),
    (26, 27), (28, 28), (29, 29), (30, 30), (31, 31),
    (32, 32), (33, 34), (35, 35), (36, 36), (37, 38),
)
SLOTS = len(DAY_SLOTS)                   # 25: five weeks of five days

NO_COL, NAME_COL = 1, 3                  # A:B and C:E
ABSENT_COL, PRESENT_COL = 39, 41         # AM:AN and AO:AP
REMARKS_COL = 43                         # AQ:AU

COLUMN_WIDTHS = {
    1: 1.68, 2: 1.34, 3: 14.11, 4: 8.73, 5: 0.17, 6: 0.84, 7: 2.02,
    8: 2.86, 9: 2.86, 10: 2.86, 11: 2.86, 12: 2.52, 13: 0.34,
    14: 2.86, 15: 2.86, 16: 2.86, 17: 2.86, 18: 0.84, 19: 2.02,
    20: 2.86, 21: 2.86, 22: 0.67, 23: 2.18, 24: 1.18, 25: 1.68,
    26: 1.68, 27: 1.18, 28: 2.86, 29: 2.86, 30: 2.86, 31: 2.86,
    32: 2.86, 33: 2.69, 34: 0.17, 35: 2.86, 36: 2.86, 37: 1.18,
    38: 1.68, 39: 1.68, 40: 5.04, 41: 5.04, 42: 1.68, 43: 10.08,
    44: 3.36, 45: 3.36, 46: 6.72, 47: 0.34,
}

HEADER_ROWS = 7                          # rows 1-7; the first learner is row 8
FIRST_LEARNER_ROW = HEADER_ROWS + 1

# --- styles ------------------------------------------------------------------

TITLE = Font(name="Arial", size=12, bold=True)
SUBTITLE = Font(name="Arial", size=8, italic=True)
LABEL = Font(name="Arial", size=9)
VALUE = Font(name="Arial", size=9, bold=True)
HEAD = Font(name="Arial", size=8, bold=True)
NAME = Font(name="Arial", size=9)
MARK = Font(name="Arial", size=9, bold=True)
TOTALS = Font(name="Arial", size=8, bold=True)
FINE = Font(name="Arial", size=7)
FINE_BOLD = Font(name="Arial", size=7, bold=True)
SIGNATURE = Font(name="Arial", size=9, bold=True)

CENTRE = Alignment(horizontal="center", vertical="center")
CENTRE_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
LEFT_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

# The form's own code for tardiness is a half-shaded cell -- upper half for a late
# comer, lower half for cutting classes. A spreadsheet cell cannot be half filled, and
# TRACKIFY only ever detects the upper case (it sees an arrival time, never a walkout),
# so the whole cell is shaded and the legend says which half it stands for.
LATE = PatternFill("solid", fgColor="BFBFBF")
HAIR = Side(style="thin", color="000000")
BOX = Border(left=HAIR, right=HAIR, top=HAIR, bottom=HAIR)

ABSENT_MARK = "x"

MALE, FEMALE, UNRECORDED = "M", "F", None

BLOCK_LABELS = {
    MALE: "<=== MALE | TOTAL Per Day ===>",
    FEMALE: "<=== FEMALE | TOTAL Per Day ===>",
    UNRECORDED: "<=== SEX NOT RECORDED | TOTAL Per Day ===>",
}

# Counted as attending. Mirrors corrections.PRESENT_STATUSES deliberately: if the two
# ever disagree, the SF2 and the register would report different numbers for the same
# month out of the same database.
PRESENT_STATUSES = ("present", "late", "online")

CONSECUTIVE_ABSENCE_ALERT = 5

# scripts/seed_demo.py seeds the users table with role placeholders rather than people,
# because sections.adviser_id needs a row to point at and there are no logins yet. A
# placeholder over a signature line reads as somebody's name, so it prints as blank and
# the adviser writes theirs in. Set [sf2] adviser_name in config.toml to fill it.
PLACEHOLDER_NAMES = ("Operator", "Class Adviser", "Administrator")

# The form's own legend, transcribed. TRACKIFY's departure from it is the second line.
CODES = ("(blank) - Present; (x)- Absent; Tardy (half shaded= Upper for Late Commer, "
         "Lower for Cutting Classes)")
CODES_NOTE = ("Generated by TRACKIFY. A shaded cell is a late arrival. Excused days "
              "are blank and named in REMARKS, not counted absent.")

REASONS = (
    ("2. REASONS/CAUSES FOR NLS", None),
    ("a. Domestic-Related Factors",
     "a.1. Had to take care of siblings\na.2. Early marriage/pregnancy\n"
     "a.3. Parents' attitude toward schooling\na.4. Family problems"),
    ("b. Individual-Related Factors",
     "b.1. Illness\nb.2. Overage\nb.3. Death\nb.4. Drug Abuse\n"
     "b.5. Poor academic performance\nb.6. Lack of interest/Distractions\n"
     "b.7. Hunger/Malnutrition"),
    ("c. School-Related Factors",
     "c.1. Teacher Factor\nc.2. Physical condition of classroom\nc.3. Peer influence"),
    ("d. Geographic/Environmental",
     "d.1. Distance between home and school\n"
     "d.2. Armed conflict (incl. Tribal wars & clanfeuds)\nd.3. Calamities/Disasters"),
    ("e. Financial-Related", "e.1. Child labor, work"),
    ("f. Others (Specify)", None),
)

GUIDELINES = (
    "1. The attendance shall be accomplished daily. Refer to the codes for checking "
    "learners' attendance.\n"
    "2. Dates shall be written in the columns after Learner's Name.\n"
    "3. To compute the following:\n"
    "     a. Percentage of Enrolment = (Registered Learners as of end of the month / "
    "Enrolment as of 1st Friday of the school year) x 100\n"
    "     b. Average Daily Attendance = Total Daily Attendance / Number of School Days "
    "in reporting month\n"
    "     c. Percentage of Attendance for the month = (Average daily attendance / "
    "Registered Learners as of end of the month) x 100\n"
    "4. Every end of the month, the class adviser will submit this form to the office "
    "of the principal for recording of summary table into School Form 4. Once signed "
    "by the principal, this form should be returned to the adviser.\n"
    "5. The adviser will provide neccessary interventions including but not limited to "
    "home visitation to learner/s who were absent for 5 consecutive days and/or those "
    "at risk of dropping out.\n"
    "6. Attendance performance of learners will be reflected in Form 137 and Form 138 "
    "every grading period.\n"
    "*Beginning of School Year cut-off report is every 1st Friday of the School Year"
)


class Sf2Error(RuntimeError):
    pass


@dataclass
class Learner:
    student_id: int
    name: str
    sex: str | None
    marks: dict[str, str] = field(default_factory=dict)      # date -> status or ""
    remarks: list[str] = field(default_factory=list)

    @property
    def present(self) -> int:
        return sum(1 for s in self.marks.values() if s in PRESENT_STATUSES)

    @property
    def absent(self) -> int:
        return sum(1 for s in self.marks.values() if s == "absent")


# --- data --------------------------------------------------------------------

def class_days(conn: sqlite3.Connection, section_id: int, year: int,
               month: int) -> list[str]:
    """The dates that get a column, in order.

    Not simply "the weekdays of the month", and school_days alone will not do either.
    sessions.get_school_day() inserts is_school_day = 1 for any date the kiosk so much
    as ticks over -- an evening when somebody opened the app creates a school day that
    nobody attended. A column for it would put "no attendance recorded" against every
    child in the section and drag the daily average down by a whole day's divisor.

    So a date has to be ATTESTED by attendance somewhere in the database, and then:

      suspended (is_school_day = 0)         never a column
      this section has attendance           a column -- it was here
      weekday, in school_days, attested     a column -- the school ran and THIS
                                            section's records are the thing missing,
                                            which is worth showing rather than hiding
      anything else                         no column
    """
    total = monthrange(year, month)[1]
    first = Date(year, month, 1).isoformat()
    last = Date(year, month, total).isoformat()

    known = {
        row["date"]: bool(row["is_school_day"])
        for row in conn.execute(
            "SELECT date, is_school_day FROM school_days WHERE date BETWEEN ? AND ?",
            (first, last),
        )
    }
    recorded = {
        row["date"] for row in conn.execute(
            """SELECT DISTINCT a.date FROM attendance_days a
               JOIN students s ON s.id = a.student_id
               WHERE s.section_id = ? AND a.date BETWEEN ? AND ?
                 AND a.superseded_by IS NULL""",
            (section_id, first, last),
        )
    }
    attested = {
        row["date"] for row in conn.execute(
            """SELECT DISTINCT date FROM attendance_days
               WHERE date BETWEEN ? AND ? AND superseded_by IS NULL""",
            (first, last),
        )
    }

    days = []
    for number in range(1, total + 1):
        day = Date(year, month, number)
        key = day.isoformat()
        if known.get(key) is False:                 # explicitly suspended
            continue
        if key in recorded:
            days.append(key)
        elif day.weekday() < 5 and key in known and key in attested:
            days.append(key)

    if len(days) > SLOTS:
        raise Sf2Error(
            f"{MONTHS[month - 1]} {year} has {len(days)} class days and School Form 2 "
            f"has room for {SLOTS}. Check for scans recorded on non-class days, or "
            "suspend those dates, before exporting."
        )
    return days


def learners(conn: sqlite3.Connection, section_id: int,
             days: list[str]) -> list[Learner]:
    """One Learner per active student, marks filled, remarks composed.

    Sorted by surname, which is the order within each block; the caller splits on .sex
    and the order survives the split.
    """
    rows = conn.execute(
        """SELECT id, first_name, last_name, sex FROM students
           WHERE section_id = ? AND active = 1
           ORDER BY last_name, first_name""",
        (section_id,),
    ).fetchall()

    records: dict[tuple[int, str], sqlite3.Row] = {}
    if days:
        placeholders = ",".join("?" * len(days))
        for row in conn.execute(
            f"""SELECT a.student_id, a.date, a.status, a.correction_reason
                FROM attendance_days a
                JOIN students s ON s.id = a.student_id
                WHERE s.section_id = ? AND a.date IN ({placeholders})
                  AND a.superseded_by IS NULL""",
            (section_id, *days),
        ):
            records[(row["student_id"], row["date"])] = row

    out = []
    for row in rows:
        learner = Learner(student_id=row["id"], sex=row["sex"],
                          name=f"{row['last_name']}, {row['first_name']}")
        excused: list[tuple[str, str]] = []
        online: list[str] = []
        missing: list[str] = []

        for day in days:
            record = records.get((row["id"], day))
            status = record["status"] if record else ""
            learner.marks[day] = status
            number = str(Date.fromisoformat(day).day)
            if record is None:
                missing.append(number)
            elif status == "excused":
                excused.append((number, (record["correction_reason"] or "").strip()))
            elif status == "online":
                online.append(number)

        if excused:
            reasons = sorted({reason for _, reason in excused if reason})
            detail = f" ({'; '.join(reasons)})" if reasons else ""
            learner.remarks.append(
                "Excused, not counted absent: "
                + ", ".join(number for number, _ in excused) + detail
            )
        if online:
            learner.remarks.append("Counted present online: " + ", ".join(online))
        if missing:
            # Left blank on the form, which SF2 reads as present -- so it has to be
            # said out loud. A blank meaning "nobody recorded this day" is not the same
            # claim as a blank meaning "this child was here".
            learner.remarks.append("No attendance recorded: " + ", ".join(missing))
        out.append(learner)
    return out


def consecutive_absences(learner: Learner, days: list[str]) -> int:
    """The longest run of absences over consecutive CLASS days.

    Consecutive in school terms, not calendar terms: Friday and the following Monday
    are consecutive, and a suspended Wednesday does not break a run because it was
    never a day the child could have attended.
    """
    longest = run = 0
    for day in days:
        run = run + 1 if learner.marks.get(day) == "absent" else 0
        longest = max(longest, run)
    return longest


# --- writing -----------------------------------------------------------------

def _put(sheet, row: int, column: int, value=None, *, font=None, align=None,
         span: tuple[int, int] | None = None, fill=None, border=True):
    """Write one cell, merging first when it spans columns.

    Merging after writing loses the value in some readers, and styling only a merge's
    top-left leaves the rest of it unbordered -- so both are handled here rather than
    at each of the couple of hundred call sites.
    """
    first, last = span or (column, column)
    if last > first:
        sheet.merge_cells(start_row=row, start_column=first,
                          end_row=row, end_column=last)
    cell = sheet.cell(row, first)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    for index in range(first, last + 1):
        target = sheet.cell(row, index)
        if border:
            target.border = BOX
        if fill:
            target.fill = fill
    return cell


def _merge(sheet, first_row: int, last_row: int, first_col: int, last_col: int,
           value=None, *, font=None, align=None, border=False):
    """A block spanning rows as well as columns -- the header band and the footer."""
    if (last_row, last_col) != (first_row, first_col):
        sheet.merge_cells(start_row=first_row, start_column=first_col,
                          end_row=last_row, end_column=last_col)
    cell = sheet.cell(first_row, first_col)
    if value is not None:
        cell.value = value
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        for row in range(first_row, last_row + 1):
            for column in range(first_col, last_col + 1):
                sheet.cell(row, column).border = BOX
    return cell


def export_sf2(
    conn: sqlite3.Connection,
    section_id: int,
    year: int,
    month: int,
    path: str | Path,
    *,
    config=None,
    school_name: str = "",
) -> Path:
    """Write one section's month as a DepEd SF2. Returns the path written."""
    path = Path(path)
    sf2 = getattr(config, "sf2", None)

    section = conn.execute(
        """SELECT sec.name, sec.grade_level, u.full_name AS adviser
           FROM sections sec LEFT JOIN users u ON u.id = sec.adviser_id
           WHERE sec.id = ?""",
        (section_id,),
    ).fetchone()
    if section is None:
        raise Sf2Error(f"No section with id {section_id}.")

    days = class_days(conn, section_id, year, month)
    people = learners(conn, section_id, days)
    blocks = [(sex, [p for p in people if p.sex == sex])
              for sex in (MALE, FEMALE, UNRECORDED)]
    # The third block exists only while somebody's sex is unrecorded. On a finished
    # roster it is absent and the form is exactly the two blocks DepEd expects.
    blocks = [(sex, members) for sex, members in blocks
              if members or sex is not UNRECORDED]

    book = Workbook()
    sheet = book.active
    sheet.title = "school_form_2"

    for column, width in COLUMN_WIDTHS.items():
        sheet.column_dimensions[get_column_letter(column)].width = width

    _titles(sheet, section, month, sf2, school_name)
    _day_header(sheet, days)
    row = _blocks(sheet, blocks, days)
    _footer(sheet, row, blocks, days, month, sf2, section)

    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    for side in ("left", "right", "top", "bottom"):
        setattr(sheet.page_margins, side, 0.2777777777777778)
    sheet.print_area = f"A1:{get_column_letter(LAST_COLUMN)}{sheet.max_row}"

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return path


def _titles(sheet, section, month, sf2, school_name) -> None:
    _merge(sheet, 1, 1, 1, LAST_COLUMN,
           "School Form 2 (SF2) Daily Attendance Report of Learners",
           font=TITLE, align=CENTRE)
    _merge(sheet, 2, 2, 1, LAST_COLUMN,
           "(This replaces Form 1, Form 2 & STS Form 4 - Absenteeism and Dropout "
           "Profile)", font=SUBTITLE, align=CENTRE)

    _merge(sheet, 3, 3, 1, 4, "School ID ", font=LABEL, align=LEFT)
    _merge(sheet, 3, 3, 6, 9, getattr(sf2, "school_id", "") or "",
           font=VALUE, align=CENTRE)
    _merge(sheet, 3, 3, 10, 12, "School Year ", font=LABEL, align=LEFT)
    _merge(sheet, 3, 3, 13, 18, getattr(sf2, "school_year", "") or "",
           font=VALUE, align=CENTRE)
    _merge(sheet, 3, 3, 19, 26, "Report for the Month of", font=LABEL, align=LEFT)
    _merge(sheet, 3, 3, 27, 33, MONTHS[month - 1], font=VALUE, align=CENTRE)

    _merge(sheet, 4, 4, 1, 4, "Name of School", font=LABEL, align=LEFT)
    _merge(sheet, 4, 4, 6, 18, school_name or "", font=VALUE, align=CENTRE)
    _merge(sheet, 4, 4, 19, 26, "Grade Level", font=LABEL, align=LEFT)
    _merge(sheet, 4, 4, 27, 33, f"Grade {section['grade_level']}",
           font=VALUE, align=CENTRE)
    _merge(sheet, 4, 4, 35, 38, "Section ", font=LABEL, align=LEFT)
    _merge(sheet, 4, 4, 39, LAST_COLUMN, section["name"], font=VALUE, align=CENTRE)

    for number, height in ((1, 30), (2, 20), (3, 20), (4, 20),
                           (5, 10), (6, 15), (7, 15)):
        sheet.row_dimensions[number].height = height


def _day_header(sheet, days: list[str]) -> None:
    """Rows 5-7: the fixed captions, then the date numbers over their weekday letters."""
    for row in (5, 6, 7):
        for column in range(1, LAST_COLUMN + 1):
            sheet.cell(row, column).border = BOX

    _merge(sheet, 5, 7, 1, 2, "No.", font=HEAD, align=CENTRE_WRAP)
    _merge(sheet, 5, 7, 3, 5, "NAME\n(Last Name, First Name, Middle Name)",
           font=HEAD, align=CENTRE_WRAP)
    _merge(sheet, 5, 5, 6, 38, "(1st row for date)", font=HEAD, align=CENTRE)
    _merge(sheet, 5, 6, 39, 42, "Total for the Month", font=HEAD, align=CENTRE_WRAP)
    _merge(sheet, 5, 7, REMARKS_COL, LAST_COLUMN,
           "REMARKS (If DROPPED OUT, state reason, please refer to legend number 2. "
           "If TRANSFERRED IN/OUT, write the name of School.)",
           font=FINE, align=CENTRE_WRAP)

    _put(sheet, 7, ABSENT_COL, "ABSENT", font=HEAD, align=CENTRE,
         span=(ABSENT_COL, ABSENT_COL + 1))
    _put(sheet, 7, PRESENT_COL, "PRESENT", font=HEAD, align=CENTRE,
         span=(PRESENT_COL, PRESENT_COL + 1))

    letters = ("M", "T", "W", "TH", "F", "SA", "SU")
    for index, (first, last) in enumerate(DAY_SLOTS):
        day = Date.fromisoformat(days[index]) if index < len(days) else None
        _put(sheet, 6, first, day.day if day else None,
             font=HEAD, align=CENTRE, span=(first, last))
        _put(sheet, 7, first, letters[day.weekday()] if day else None,
             font=HEAD, align=CENTRE, span=(first, last))


def _blocks(sheet, blocks, days: list[str]) -> int:
    """Write every learner block and its total row. Returns the row after the last."""
    row = FIRST_LEARNER_ROW
    running: list[list[int]] = []

    for sex, members in blocks:
        totals = [0] * SLOTS
        for number, learner in enumerate(members, start=1):
            sheet.row_dimensions[row].height = 20
            _put(sheet, row, NO_COL, number, font=NAME, align=CENTRE, span=(1, 2))
            _put(sheet, row, NAME_COL, learner.name, font=NAME, align=LEFT, span=(3, 5))

            for index, (first, last) in enumerate(DAY_SLOTS):
                status = learner.marks.get(days[index]) if index < len(days) else None
                _put(sheet, row, first,
                     ABSENT_MARK if status == "absent" else None,
                     font=MARK, align=CENTRE, span=(first, last),
                     fill=LATE if status == "late" else None)
                if status in PRESENT_STATUSES:
                    totals[index] += 1

            _put(sheet, row, ABSENT_COL, learner.absent, font=NAME, align=CENTRE,
                 span=(ABSENT_COL, ABSENT_COL + 1))
            _put(sheet, row, PRESENT_COL, learner.present, font=NAME, align=CENTRE,
                 span=(PRESENT_COL, PRESENT_COL + 1))
            _put(sheet, row, REMARKS_COL, "; ".join(learner.remarks) or None,
                 font=FINE, align=LEFT_WRAP, span=(REMARKS_COL, LAST_COLUMN))
            row += 1

        running.append(totals)
        row = _total_row(sheet, row, BLOCK_LABELS[sex], len(members), totals, days)

    combined = [sum(column) for column in zip(*running)] if running else [0] * SLOTS
    row = _total_row(sheet, row, "Combined TOTAL Per Day",
                     sum(len(members) for _, members in blocks), combined, days)
    return row


def _total_row(sheet, row: int, label: str, count: int, totals: list[int],
               days: list[str]) -> int:
    sheet.row_dimensions[row].height = 20
    _put(sheet, row, NO_COL, count, font=TOTALS, align=CENTRE, span=(1, 2))
    _put(sheet, row, NAME_COL, label, font=TOTALS, align=CENTRE, span=(3, 5))
    for index, (first, last) in enumerate(DAY_SLOTS):
        _put(sheet, row, first, totals[index] if index < len(days) else None,
             font=TOTALS, align=CENTRE, span=(first, last))
    _put(sheet, row, ABSENT_COL, None, span=(ABSENT_COL, ABSENT_COL + 1))
    _put(sheet, row, PRESENT_COL, None, span=(PRESENT_COL, PRESENT_COL + 1))
    _put(sheet, row, REMARKS_COL, None, span=(REMARKS_COL, LAST_COLUMN))
    return row + 1


# --- the footer: guidelines, legend, summary, signatures ---------------------

def summary(blocks, days: list[str]) -> dict:
    """The right-hand summary panel, computed.

    Only the rows TRACKIFY can actually answer. Enrolment as of the first Friday,
    transfers and drop-outs have no column in this database, so they are absent from
    this dict and print as blank boxes -- as does the percentage of enrolment, which is
    a ratio against a figure nobody here holds.
    """
    by_sex = {sex: members for sex, members in blocks}
    male = by_sex.get(MALE, [])
    female = by_sex.get(FEMALE, [])
    everyone = [person for members in by_sex.values() for person in members]

    def daily_average(members) -> float | None:
        if not days:
            return None
        return sum(person.present for person in members) / len(days)

    registered = {"M": len(male), "F": len(female), "total": len(everyone)}
    average = {
        "M": daily_average(male), "F": daily_average(female),
        "total": daily_average(everyone),
    }
    percentage = None
    if average["total"] is not None and registered["total"]:
        percentage = average["total"] / registered["total"] * 100

    return {
        "days": len(days),
        "registered": registered,
        "average": average,
        "percentage": percentage,
        "five_day_absentees": sum(
            1 for person in everyone
            if consecutive_absences(person, days) >= CONSECUTIVE_ABSENCE_ALERT
        ),
        "unrecorded_sex": len(by_sex.get(UNRECORDED, [])),
    }


def _footer(sheet, row: int, blocks, days: list[str], month: int, sf2,
            section) -> None:
    """Everything below the combined-total row. Laid out relative to `row`."""
    stats = summary(blocks, days)
    top = row + 1                      # one blank row, as on the form

    # -- left: the guidelines slab -------------------------------------------
    _merge(sheet, top, top, 1, 24, "GUIDELINES:", font=FINE_BOLD, align=LEFT)
    _merge(sheet, top + 1, top + 17, 1, 24, GUIDELINES, font=FINE, align=LEFT_WRAP)

    # -- middle: the codes and the NLS reasons -------------------------------
    _merge(sheet, top, top, 26, 37, "1. CODES FOR CHECKING ATTENDANCE",
           font=FINE_BOLD, align=LEFT)
    _merge(sheet, top + 1, top + 1, 26, 37, CODES, font=FINE, align=LEFT_WRAP)
    _merge(sheet, top + 2, top + 2, 26, 37, CODES_NOTE, font=FINE, align=LEFT_WRAP)

    line = top + 3
    for heading, detail in REASONS:
        _merge(sheet, line, line, 26, 37, heading, font=FINE_BOLD, align=LEFT)
        line += 1
        if detail:
            height = detail.count("\n") + 1
            _merge(sheet, line, line + height - 1, 26, 37, detail,
                   font=FINE, align=LEFT_WRAP)
            line += height
    _merge(sheet, line, line, 26, 37, "Generated thru TRACKIFY", font=FINE, align=LEFT)

    # -- right: the summary panel --------------------------------------------
    _summary_panel(sheet, top, stats, month)
    _signatures(sheet, top + 23, sf2, section)


def _summary_panel(sheet, top: int, stats: dict, month: int) -> None:
    """The boxed table at AM:AU. A row with no value prints an empty box on purpose."""
    _merge(sheet, top, top, 39, 41, f"Month : {MONTHS[month - 1]}",
           font=FINE_BOLD, align=LEFT)
    _merge(sheet, top, top, 42, 43,
           f"No. of Days of Classes: {stats['days']}", font=FINE_BOLD, align=LEFT)
    _merge(sheet, top, top, 44, LAST_COLUMN, "Summary", font=FINE_BOLD, align=CENTRE,
           border=True)

    header = top + 1
    for column, label in ((44, "M"), (45, "F"), (46, "TOTAL")):
        _merge(sheet, header, header, column,
               LAST_COLUMN if column == 46 else column,
               label, font=FINE_BOLD, align=CENTRE, border=True)

    # (label, values or None). None leaves the three boxes empty -- see summary().
    lines = (
        ("* Enrolment as of (1st Friday of June)", None),
        ("Late enrolment during the month (beyond cut-off)", None),
        ("Registered Learners as of end of month", stats["registered"]),
        ("Percentage of Enrolment as of end of month", None),
        ("Average Daily Attendance", stats["average"]),
        ("Percentage of Attendance for the month",
         {"total": stats["percentage"]}),
        ("Number of students absent for 5 consecutive days",
         {"total": stats["five_day_absentees"]}),
        ("Dropped out", None),
        ("Transferred out", None),
        ("Transferred in", None),
    )

    line = header + 1
    for label, values in lines:
        _merge(sheet, line, line, 39, 43, label, font=FINE, align=LEFT, border=True)
        for column, key in ((44, "M"), (45, "F"), (46, "total")):
            last = LAST_COLUMN if key == "total" else column
            value = (values or {}).get(key)
            if isinstance(value, float):
                value = round(value, 2)
            _merge(sheet, line, line, column, last, value,
                   font=FINE, align=CENTRE, border=True)
        line += 1

    if stats["unrecorded_sex"]:
        _merge(sheet, line, line, 39, LAST_COLUMN,
               f"{stats['unrecorded_sex']} learner(s) have no sex recorded and are "
               "listed in a third block. Set it in the student roster; this form is "
               "not a valid SF2 until they are placed.",
               font=FINE, align=LEFT_WRAP)


def _signatures(sheet, top: int, sf2, section) -> None:
    adviser = (getattr(sf2, "adviser_name", "") or "").strip()
    if not adviser:
        # The section's own adviser is the right person -- when the users table holds a
        # person. See PLACEHOLDER_NAMES.
        found = (section["adviser"] or "").strip()
        adviser = "" if found in PLACEHOLDER_NAMES else found
    head = (getattr(sf2, "school_head_name", "") or "").strip()

    _merge(sheet, top, top, 39, LAST_COLUMN,
           "I certify that this is a true and correct report.", font=FINE, align=LEFT)
    _merge(sheet, top + 2, top + 2, 40, LAST_COLUMN, adviser or None,
           font=SIGNATURE, align=CENTRE)
    _merge(sheet, top + 3, top + 3, 40, LAST_COLUMN,
           "(Signature of Adviser over Printed Name)", font=FINE, align=CENTRE)
    _merge(sheet, top + 5, top + 5, 39, LAST_COLUMN, "Attested by:",
           font=FINE, align=LEFT)
    _merge(sheet, top + 7, top + 7, 40, LAST_COLUMN, head or None,
           font=SIGNATURE, align=CENTRE)
    _merge(sheet, top + 8, top + 8, 40, LAST_COLUMN,
           "(Signature of School Head over Printed Name)", font=FINE, align=CENTRE)

    for offset in (2, 7):
        for column in range(40, LAST_COLUMN + 1):
            sheet.cell(top + offset, column).border = Border(bottom=HAIR)


def default_filename(label: str, year: int, month: int) -> str:
    safe = "".join(c if c.isalnum() or c in "-_" else "-" for c in label)
    return f"SF2-{safe}-{year}-{month:02d}.xlsx"
