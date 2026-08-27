"""Analytics export: trend, risk, AHP weights, screening, model quality.

One workbook, one sheet per topic, so a supervisor or a judge can open a single file and
read the whole analysis. The attendance register stays in its own SF2-shaped export --
that is the form staff already check, and burying it in a report would stop it being
checked.

**Every sheet is written even when its figure could not be computed**, and says why in
its first cell. An absent sheet reads as a bug and a zero reads as a finding; "a trend
needs at least 3 school days and this database has 1" reads as an instruction. With no
scanning done yet that path is the normal case, not an edge case.
"""

from __future__ import annotations

import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..analytics import ahp, risk, screening, trend
from . import safe_filename

HEAD = Font(bold=True, size=11)
TITLE = Font(bold=True, size=14)
SMALL = Font(size=9, italic=True, color="666666")
WARN = Font(bold=True, size=10, color="9C5700")
CENTRE = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(wrap_text=True, vertical="top")

CAUTION = PatternFill("solid", fgColor="FFF3CD")
HEADER_FILL = PatternFill("solid", fgColor="EDF1F0")
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

BAND_FILL = {
    "Low": PatternFill("solid", fgColor="E8F5E9"),
    "Monitor": PatternFill("solid", fgColor="FFF8E1"),
    "Elevated": PatternFill("solid", fgColor="FFE0B2"),
    "High": PatternFill("solid", fgColor="FFCDD2"),
}

CRITERION_LABELS = ("Absence risk", "Tardiness", "Early departure")


def _headers(sheet, row: int, labels, widths=None) -> None:
    for index, label in enumerate(labels):
        cell = sheet.cell(row, 1 + index, label)
        cell.font = HEAD
        cell.fill = HEADER_FILL
        cell.border = BOX
        if widths:
            sheet.column_dimensions[get_column_letter(1 + index)].width = widths[index]


def _title(sheet, text: str, subtitle: str = "") -> int:
    sheet["A1"] = text
    sheet["A1"].font = TITLE
    if subtitle:
        sheet["A2"] = subtitle
        sheet["A2"].font = SMALL
        return 4
    return 3


def _blocked(sheet, title: str, reason: str) -> None:
    """A sheet that exists, is obviously empty on purpose, and says what is needed."""
    _title(sheet, title)
    cell = sheet.cell(3, 1, reason)
    cell.font = WARN
    cell.fill = CAUTION
    cell.alignment = WRAP
    sheet.merge_cells(start_row=3, start_column=1, end_row=6, end_column=6)
    sheet.column_dimensions["A"].width = 30
    for column in "BCDEF":
        sheet.column_dimensions[column].width = 16


def _note(sheet, row: int, text: str, width: int = 6) -> int:
    cell = sheet.cell(row, 1, text)
    cell.font = SMALL
    cell.alignment = WRAP
    sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    return row + 1


# --- sheets -----------------------------------------------------------------

def _summary_sheet(sheet, *, school_name, scope, period, fit, report, summary,
                   weights) -> None:
    _title(sheet, school_name or "TRACKIFY", "Analytics report")
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 62

    rows = [
        ("Generated", datetime.now().isoformat(timespec="seconds")),
        ("Scope", scope),
        ("Observation period", period),
        ("Attendance trend",
         f"slope {fit.slope:+.5f} per school day over {fit.n} days"
         if fit.ok else "not computed"),
        ("Students scored", str(len(report.rows)) if report.ok else "0"),
        ("Absence model",
         report.model.validation if report.model else "not fitted"),
        ("Risk weights",
         "elicited from " + weights.elicited_from if weights.elicited
         else "PLACEHOLDER - not elicited"),
        ("Screening coverage",
         f"{summary.coverage:.1%}" if summary.coverage is not None else "no scans yet"),
    ]
    row = 4
    for label, value in rows:
        sheet.cell(row, 1, label).font = HEAD
        sheet.cell(row, 2, value).alignment = WRAP
        row += 1

    row += 1
    sheet.cell(row, 1, "Read this first").font = HEAD
    row += 1
    caveats = []
    if weights.caveat:
        caveats.append(weights.caveat)
    if not fit.ok:
        caveats.append("Trend: " + fit.reason)
    else:
        caveats.extend(fit.caveats)
    if report.model_note:
        caveats.append("Absence model: " + report.model_note +
                       " P(absent) is each student's observed absence rate instead, "
                       "labelled as such in the Risk sheet.")
    caveats.extend(summary.notes)
    caveats.append(
        "The composite score RECOMMENDS REVIEW. It never imposes a sanction: a person "
        "decides in every case."
    )
    for text in caveats:
        cell = sheet.cell(row, 1, "- " + text)
        cell.alignment = WRAP
        sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        sheet.row_dimensions[row].height = 30
        row += 1


def _trend_sheet(sheet, fit) -> None:
    if not fit.ok:
        _blocked(sheet, "Attendance trend", fit.reason)
        return

    _title(sheet, "Attendance trend",
           "Daily attendance rate regressed on school-day index (OLS)")
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 16

    stats = [
        ("School days (n)", fit.n),
        ("Slope (b)", round(fit.slope, 6)),
        ("Intercept (a)", round(fit.intercept, 6)),
        ("R squared", round(fit.r_squared, 4)),
        ("p-value for slope", round(fit.p_value, 5)),
        ("95% CI lower", round(fit.ci_low, 6)),
        ("95% CI upper", round(fit.ci_high, 6)),
        ("Durbin-Watson", round(fit.durbin_watson, 4)),
        ("Reading", fit.direction),
    ]
    row = 4
    for label, value in stats:
        sheet.cell(row, 1, label).font = HEAD
        sheet.cell(row, 2, value)
        row += 1

    row += 1
    for text in fit.caveats:
        row = _note(sheet, row, text)
    row = _note(sheet, row,
                "The slope is the change in attendance RATE per school day. y is the "
                "daily rate, never a running total: a cumulative series fits at R^2 ~ "
                "0.99 and only restates the mean.")

    row += 1
    _headers(sheet, row, ("Date", "Attendance rate"), (30, 16))
    row += 1
    for date, rate in fit.days:
        sheet.cell(row, 1, date).border = BOX
        cell = sheet.cell(row, 2, rate)
        cell.number_format = "0.0%"
        cell.border = BOX
        cell.alignment = CENTRE
        row += 1


def _risk_sheet(sheet, report, config) -> None:
    if not report.ok:
        _blocked(sheet, "Composite risk",
                 report.model_note or "No students to score yet.")
        return

    _title(sheet, "Composite risk",
           "Risk = w_A*P(absent) + w_T*T + w_E*E, then raised to the incident floor "
           "-- recommends review, never a sanction")

    # The incident columns sit immediately after Band, which keeps the band cell at
    # column 9 for the fill below and reads in the right order: the band, then why.
    columns = ("LRN", "Student", "Section", "P(absent)", "P source", "Tardiness T",
               "Early departure E", "Composite", "Band", "Incidents", "Kind",
               "Max severity", "Band source", "Late", "Early", "Absent", "Days")
    widths = (14, 30, 15, 11, 26, 12, 17, 11, 11, 10, 20, 12, 26, 7, 7, 8, 7)
    row = 4
    _headers(sheet, row, columns, widths)
    row += 1

    for entry in report.rows:
        # Category and severity, never incidents.item_description. The schema calls a
        # record naming a minor beside a prohibited item sensitive personal information
        # under RA 10173; the free text adds nothing for triage that the category does
        # not, and this workbook gets emailed.
        values = (entry.lrn, entry.name, entry.section,
                  round(entry.p_absent, 4), entry.p_absent_source,
                  round(entry.tardiness, 4), round(entry.early_departure, 4),
                  round(entry.composite, 4), entry.band,
                  entry.n_incidents or "", ", ".join(entry.incident_kinds),
                  entry.max_severity or "", entry.band_source,
                  entry.n_late, entry.n_early, entry.n_absent, entry.n_days)
        for index, value in enumerate(values):
            cell = sheet.cell(row, 1 + index, value)
            cell.border = BOX
            if index >= 3:
                cell.alignment = CENTRE
        sheet.cell(row, 9).fill = BAND_FILL.get(entry.band, CAUTION)
        row += 1

    row += 1
    counts = report.by_band()
    sheet.cell(row, 1, "Band totals").font = HEAD
    row += 1
    thresholds = (("Low", f"below {config.risk.band_low}"),
                  ("Monitor", f"{config.risk.band_low} to {config.risk.band_monitor}"),
                  ("Elevated", f"{config.risk.band_monitor} to {config.risk.band_elevated}"),
                  ("High", f"{config.risk.band_elevated} and above"))
    for band, span in thresholds:
        sheet.cell(row, 1, band).fill = BAND_FILL[band]
        sheet.cell(row, 2, counts.get(band, 0)).alignment = CENTRE
        sheet.cell(row, 3, span).font = SMALL
        row += 1

    row += 1
    if config.risk.bands_set_by:
        when = f" on {config.risk.bands_set_on}" if config.risk.bands_set_on else ""
        row = _note(sheet, row,
                    f"Band cutoffs set by {config.risk.bands_set_by}{when}. A boundary "
                    "decides whether a real student is referred, which is an "
                    "institutional decision, not a researcher's.",
                    width=len(columns))
    else:
        row = _note(sheet, row,
                    "Band cutoffs come from config.toml and are PLACEHOLDERS until the "
                    "school sets them. A boundary decides whether a real student is "
                    "referred, which is an institutional decision, not a researcher's.",
                    width=len(columns))
    row = _note(sheet, row,
                "A confirmed prohibited-item incident sets a MINIMUM band by severity "
                "(config.toml [risk.incident_floor]). It raises a band, never lowers "
                "one, and does not change the composite -- 'Band source' says which "
                "rule decided. It is a floor rather than a weighted criterion because "
                "one incident saturates to 0.2212, below the 0.30 Monitor cutoff: no "
                "weight summing to 1 with the others could raise a band on its own.",
                width=len(columns))
    if report.model_note:
        _note(sheet, row,
              "P(absent) is the observed absence rate, not a model prediction: "
              + report.model_note, width=len(columns))


def _ahp_sheet(sheet, weights, matrix) -> None:
    _title(sheet, "AHP weights",
           "Pairwise judgements, geometric-mean weights, and the consistency check")
    sheet.column_dimensions["A"].width = 22
    for column in "BCDE":
        sheet.column_dimensions[column].width = 15

    row = 4
    if weights.caveat:
        cell = sheet.cell(row, 1, weights.caveat)
        cell.font = WARN
        cell.fill = CAUTION
        cell.alignment = WRAP
        sheet.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=5)
        sheet.row_dimensions[row].height = 30
        row += 4

    sheet.cell(row, 1, "Pairwise comparison matrix").font = HEAD
    row += 1
    _headers(sheet, row, ("", *CRITERION_LABELS))
    row += 1
    for index, label in enumerate(CRITERION_LABELS):
        sheet.cell(row, 1, label).font = HEAD
        for column, value in enumerate(matrix[index]):
            cell = sheet.cell(row, 2 + column, round(float(value), 4))
            cell.border = BOX
            cell.alignment = CENTRE
        row += 1

    row += 1
    sheet.cell(row, 1, "Derived weights").font = HEAD
    row += 1
    # as_dict() is keyed by ahp.CRITERIA, so this no longer restates the three names in
    # a fourth place -- and strict catches a labels tuple that has drifted from them.
    for label, value in zip(CRITERION_LABELS, weights.as_dict().values(), strict=True):
        sheet.cell(row, 1, label)
        cell = sheet.cell(row, 2, round(value, 4))
        cell.alignment = CENTRE
        cell.border = BOX
        row += 1
    sheet.cell(row, 1, "Sum").font = HEAD
    sheet.cell(row, 2, round(weights.absence + weights.tardiness
                             + weights.early_departure, 4)).alignment = CENTRE
    row += 2

    sheet.cell(row, 1, "Consistency check").font = HEAD
    row += 1
    for label, value in (("lambda max", round(weights.lambda_max, 4)),
                         ("Consistency Index", round(weights.ci, 4)),
                         ("Random Index (n=3)", ahp.RANDOM_INDEX[weights.n]),
                         ("Consistency Ratio", round(weights.cr, 4)),
                         ("Usable (CR <= 0.10)", "yes" if weights.consistent else "NO")):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value).alignment = CENTRE
        row += 1

    row += 1
    if weights.elicited:
        row = _note(sheet, row,
                    f"Elicited from {weights.elicited_from} on {weights.elicited_at} "
                    f"(version {weights.version}).")
    row = _note(sheet, row,
                "Three criteria, not two: a 2x2 pairwise matrix is perfectly consistent "
                "by construction, so its consistency ratio would prove nothing.")


def _screening_sheet(sheet, summary) -> None:
    _title(sheet, "Screening and incidents",
           "Descriptive only. Counts, never names -- RA 10173.")
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 14
    sheet.column_dimensions["C"].width = 46

    row = 4
    sheet.cell(row, 1, "Procedure").font = HEAD
    row += 1
    rate = lambda v: "-" if v is None else f"{v:.1%}"          # noqa: E731
    for label, value, note in (
        ("Scans", summary.scans, "in and out"),
        ("Arrivals", summary.arrivals, "screening happens on the way in"),
        ("Screenings answered", summary.screened, "including 'not screened'"),
        ("Examined", summary.examined, "someone actually looked"),
        ("Coverage", rate(summary.coverage), "examined / arrivals"),
        ("Metal detected (alarms)", summary.alarms, ""),
        ("Alarm rate", rate(summary.alarm_rate), "alarms / examined"),
        ("Confirmed findings", summary.confirmed, "prohibited or school hazard"),
        ("Confirmation rate", rate(summary.confirmation_rate), "confirmed / alarms"),
    ):
        sheet.cell(row, 1, label).font = HEAD
        sheet.cell(row, 2, value).alignment = CENTRE
        sheet.cell(row, 3, note).font = SMALL
        row += 1

    row += 1
    sheet.cell(row, 1, "Outcomes").font = HEAD
    row += 1
    for name, count in summary.outcomes.items():
        sheet.cell(row, 1, name.replace("_", " "))
        sheet.cell(row, 2, count).alignment = CENTRE
        row += 1

    row += 1
    sheet.cell(row, 1, f"Incidents by category (total {summary.incident_total})").font = HEAD
    row += 1
    for name, count in summary.incidents_by_category.items():
        sheet.cell(row, 1, name)
        sheet.cell(row, 2, count).alignment = CENTRE
        row += 1

    row += 1
    sheet.cell(row, 1, "Incidents by severity").font = HEAD
    row += 1
    for level, count in summary.incidents_by_severity.items():
        sheet.cell(row, 1, f"severity {level}")
        sheet.cell(row, 2, count).alignment = CENTRE
        row += 1
    sheet.cell(row, 1, "Severity total").font = HEAD
    sheet.cell(row, 2, summary.severity_total).alignment = CENTRE
    row += 2

    sheet.cell(row, 1, f"Custody items (total {summary.custody_total})").font = HEAD
    row += 1
    for name, count in (summary.custody_by_status or {"none recorded": 0}).items():
        sheet.cell(row, 1, str(name))
        sheet.cell(row, 2, count).alignment = CENTRE
        row += 1
    sheet.cell(row, 1, "Released with no hazard request").font = HEAD
    sheet.cell(row, 2, summary.released_unbacked).alignment = CENTRE
    row += 1
    sheet.cell(row, 1, "Hazard requests on file").font = HEAD
    sheet.cell(row, 2, summary.hazard_requests).alignment = CENTRE
    row += 2

    for text in summary.notes:
        row = _note(sheet, row, text)
    row = _note(sheet, row,
                "Incidents are NOT a weighted term in the composite. Over a short study "
                "the count is near zero for every student, and a near-constant criterion "
                "contributes noise, cannot be validated, and invites the question of how "
                "its weight was derived.")
    _note(sheet, row,
          "They are not ignored either: a confirmed incident sets a MINIMUM band on the "
          "Risk sheet, keyed to severity. A floor rather than a weight because one "
          "incident saturates to 0.2212 against a 0.30 Monitor cutoff -- a weighted term "
          "could never have raised the band at all. Counts here stay aggregate; the "
          "per-student detail is on the Risk sheet.")


def _model_sheet(sheet, report) -> None:
    if report.model is None:
        _blocked(sheet, "Absence model quality",
                 (report.model_note or "The model was not fitted.")
                 + " Until it is, P(absent) in the Risk sheet is each student's observed "
                   "absence rate, which describes the past rather than predicting the "
                   "next day.")
        return

    quality = report.model
    _title(sheet, "Absence model quality",
           "Pooled logistic regression. Precision, recall and AUC -- not accuracy.")
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 16
    sheet.column_dimensions["C"].width = 48

    row = 4
    for label, value, note in (
        ("Observations", quality.n_rows, "one row per student per school day"),
        ("Absence events", quality.n_events, ""),
        ("Event rate", f"{quality.event_rate:.1%}", ""),
        ("Validation", quality.validation, ""),
        ("Precision", round(quality.precision, 4), "of those flagged, how many were absent"),
        ("Recall", round(quality.recall, 4), "of those absent, how many were flagged"),
        ("ROC AUC", round(quality.roc_auc, 4), "0.5 is chance"),
    ):
        sheet.cell(row, 1, label).font = HEAD
        sheet.cell(row, 2, value).alignment = CENTRE
        sheet.cell(row, 3, note).font = SMALL
        row += 1

    row += 1
    sheet.cell(row, 1, "Confusion matrix").font = HEAD
    row += 1
    _headers(sheet, row, ("", "Predicted absent", "Predicted present"))
    row += 1
    for label, first, second in (("Actually absent", quality.true_pos, quality.false_neg),
                                 ("Actually present", quality.false_pos, quality.true_neg)):
        sheet.cell(row, 1, label).font = HEAD
        sheet.cell(row, 2, first).alignment = CENTRE
        sheet.cell(row, 3, second).alignment = CENTRE
        row += 1

    row += 1
    sheet.cell(row, 1, "Coefficients (log-odds)").font = HEAD
    row += 1
    sheet.cell(row, 1, "intercept")
    sheet.cell(row, 2, round(quality.intercept, 4)).alignment = CENTRE
    row += 1
    for name, value in quality.coefficients.items():
        sheet.cell(row, 1, name)
        sheet.cell(row, 2, round(value, 4)).alignment = CENTRE
        row += 1

    row += 1
    row = _note(sheet, row,
                "Accuracy is deliberately not reported: when absences are rare, always "
                "predicting 'present' scores high accuracy and is useless.")
    _note(sheet, row,
          "One pooled model with per-student features, not a model per student -- "
          "20 school days gives about 20 observations each, far too few to fit five "
          "coefficients.")


# --- entry point -------------------------------------------------------------

def export_analytics(conn: sqlite3.Connection, config, path: str | Path, *,
                     section_id: int | None = None, start: str | None = None,
                     end: str | None = None, school_name: str = "",
                     persist: bool = False) -> Path:
    """Write the analytics workbook. Returns the path written."""
    path = Path(path)

    fit = trend.attendance_trend(conn, section_id=section_id, start=start, end=end)
    report = risk.compute(conn, config, section_id=section_id, end=end, persist=persist)
    summary = screening.summarise(conn, start=start, end=end)
    weights = report.weights or ahp.active(conn)
    matrix = ahp.matrix_of(conn, weights)

    scope = "All sections"
    if section_id is not None:
        row = conn.execute(
            "SELECT grade_level, name FROM sections WHERE id = ?", (section_id,)
        ).fetchone()
        scope = f"{row['grade_level']}-{row['name']}" if row else f"section {section_id}"
    period = f"{start or 'earliest'} to {end or 'latest'}"

    book = Workbook()
    _summary_sheet(book.active, school_name=school_name, scope=scope, period=period,
                   fit=fit, report=report, summary=summary, weights=weights)
    book.active.title = "Summary"

    _trend_sheet(book.create_sheet("Trend"), fit)
    _risk_sheet(book.create_sheet("Risk"), report, config)
    _ahp_sheet(book.create_sheet("AHP"), weights, matrix)
    _screening_sheet(book.create_sheet("Screening"), summary)
    _model_sheet(book.create_sheet("Model"), report)

    path.parent.mkdir(parents=True, exist_ok=True)
    book.save(path)
    return path


def default_filename(scope: str) -> str:
    return f"trackify-analytics-{safe_filename(scope)}-{datetime.now():%Y%m%d}.xlsx"
