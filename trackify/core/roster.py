"""Read a real roster out of the school's spreadsheet.

The school keeps its class lists as one worksheet per section, with columns for LRN,
student name, parent name, parent mobile and parent email. Roughly half the rows are
incomplete -- a class list is a living document and the office fills it in over weeks --
so importing has to filter, and has to say what it filtered and why. A silent import
that seeds 72 of 124 students looks like a bug on the morning of the fair.

The split between a *rejection* and a *note* is the whole design:

  rejected  the record cannot exist -- no LRN, or nobody to notify
  noted     the record exists but something about it looks wrong

Nothing here corrects anything. An LRN that is one digit too long is almost certainly a
typo, but writing a guessed LRN into a DepEd-shaped record is worse than carrying a
flagged one, so the odd value is stored verbatim and reported.

Parsing lives here rather than in scripts/seed_demo.py so the rules can be tested against
synthetic rows. The real workbook holds 124 real children's records and no test should
open it.
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass, field
from pathlib import Path

from .mobile import InvalidMobile, normalise

LRN_LENGTH = 12

# Section banners inside the sheet -- the office splits each list into a male and a female
# block, so these appear mid-sheet as well as at the top. Matched by content, never by row
# number: the banner sits at row 19 in one sheet and row 24 in the others.
BANNERS = {"MALE", "FEMALE"}
HEADER_FIRST_CELL = "LRN"

_SECTION = re.compile(r"^\s*(\d+)\s*[-–]\s*(.+?)\s*$")


@dataclass(frozen=True)
class Candidate:
    """A student the importer will seed."""

    lrn: str
    first: str
    last: str
    section_name: str
    grade_level: int
    guardian_name: str
    # None means the sheet had a number but it could not be parsed. That is different
    # from the field being blank, which is a rejection -- see reject_reasons().
    guardian_mobile: str | None
    notes: tuple[str, ...] = ()

    @property
    def full_name(self) -> str:
        return f"{self.last}, {self.first}"

    @property
    def section_label(self) -> str:
        return f"{self.grade_level}-{self.section_name}"


@dataclass(frozen=True)
class Rejected:
    """A row that cannot become a student, and why."""

    name: str
    section_label: str
    reasons: tuple[str, ...] = field(default=())


def cell(value: object) -> str:
    """Normalise one spreadsheet cell to clean text.

    openpyxl hands back a float for anything Excel stored as a number, so an LRN arrives
    as 111995150037.0 and a mobile as 9478179371.0. str() on those keeps the '.0' and
    turns a valid number into a rejected one, which is a silent and very confusing loss.
    """
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def parse_section(title: str) -> tuple[int, str]:
    """'11-Initiative' -> (11, 'Initiative').

    Kept as (grade_level, name) because that is how the sections table stores it, and the
    rest of the application rebuilds the label with f"{grade_level}-{name}".
    """
    match = _SECTION.match(title)
    if not match:
        # A sheet named without a grade prefix still has to import. Grade 0 sorts first
        # and is obviously wrong on screen, which is better than refusing the sheet.
        return 0, title.strip()
    return int(match.group(1)), match.group(2)


def split_name(raw: str) -> tuple[str, str]:
    """'Ellorenco, Marlou Alan Jr. B.' -> ('Ellorenco', 'Marlou Alan Jr. B.').

    Split on the FIRST comma only. Suffixes and middle initials stay with the given name:
    students table has no middle-name column, and dropping the initial would discard
    something the school actually recorded.
    """
    last, _, first = raw.partition(",")
    return last.strip(), first.strip()


def is_skippable(lrn: str, name: str) -> bool:
    """True for the MALE/FEMALE banners and the column-header row."""
    if not name and not lrn:
        return True
    if lrn.rstrip(":").strip().upper() == HEADER_FIRST_CELL:
        return True
    if name.upper().startswith("NAME OF"):
        return True
    banner = (name or lrn).rstrip(":").strip().upper()
    return banner in BANNERS


def lrn_note(lrn: str) -> str | None:
    """Flag an LRN that is not 12 digits -- without touching it.

    A DepEd LRN is 12 digits. Five students in the real sheet have 11 or 13, each a
    plausible single-digit slip. They import as written; somebody who knows the child
    fixes the spreadsheet.
    """
    if not lrn.isdigit():
        return f"LRN {lrn!r} contains non-digits - no QR code can be made for it"
    if lrn.startswith("0"):
        # A payload is built by encode(int(lrn)), so a leading zero is lost and the card
        # would never resolve back to this student. Caught here, at import, rather than
        # as an unexplained "Student not found" at the gate on the morning it matters.
        return (f"LRN {lrn} starts with 0 - it cannot survive a QR payload, "
                "so this student's card would never scan")
    if len(lrn) != LRN_LENGTH:
        return f"LRN {lrn} is {len(lrn)} digits, expected {LRN_LENGTH}"
    return None


def _fold(text: str) -> str:
    """Strip accents for comparison only. Never for storage -- a name is spelled the way
    its owner spells it, and 'Pena' is a different name from 'Peña'."""
    stripped = unicodedata.normalize("NFD", text)
    return "".join(c for c in stripped if not unicodedata.combining(c)).casefold()


def surname_note(last: str, guardian: str) -> str | None:
    """Flag a surname that matches the parent's except for its diacritics.

    One real case in the sheet: the student is typed 'Seňora' (n with caron, U+0148) and
    the parent 'Senora' (n with tilde, U+00F1). Two characters that look nearly identical
    at 11pt and sort differently everywhere. Flagged, never corrected -- guessing which
    spelling a family uses is not the importer's call.
    """
    if not last or not guardian:
        return None
    guardian_last = guardian.partition(",")[0].strip()
    # Only compare when the parent's name is written surname-first; 'Maria Cruz' would
    # otherwise compare a given name against a surname and flag every second row.
    if "," not in guardian or not guardian_last:
        return None
    if last == guardian_last or _fold(last) != _fold(guardian_last):
        return None
    return (f"surname is spelled {last!r} but the parent's is {guardian_last!r} "
            "- same name, different characters")


def parse_row(row: tuple, grade_level: int, section_name: str) -> Candidate | Rejected | None:
    """One spreadsheet row -> a student, a rejection, or None for a banner/header."""
    lrn, name, guardian, mobile = (cell(v) for v in _first_four(row))

    if is_skippable(lrn, name):
        return None

    label = f"{grade_level}-{section_name}"
    reasons: list[str] = []

    if not name:
        # An LRN with no name beside it. is_skippable() has already absorbed the truly
        # blank rows, so this is a real anomaly and belongs in the report rather than
        # disappearing -- somebody typed an LRN for a child whose name went missing.
        return Rejected(name=f"(unnamed, LRN {lrn})", section_label=label,
                        reasons=("no student name",))
    if not lrn:
        # The only hard requirement. An LRN is the student's identity here: it is what a
        # QR card is signed over, so a student without one cannot be scanned at all and
        # there is nothing for a later import to match them against.
        reasons.append("no LRN")

    if reasons:
        return Rejected(name=name, section_label=label, reasons=tuple(reasons))

    notes: list[str] = []
    if note := lrn_note(lrn):
        notes.append(note)

    # Guardian details are OPTIONAL, and used to be fatal. They are not, because the
    # roster UI exists precisely so staff can fill them in afterwards -- refusing the
    # student here would mean the only way to fix an empty contact column is Excel,
    # which is the thing the UI replaces. It also keeps this rule identical to the QR
    # generator's, and the two disagreeing is what produced a box of printed cards for
    # students who were never in the database.
    if not guardian:
        notes.append("no parent name on file")
    if not mobile:
        notes.append("no parent mobile on file - this guardian cannot be texted")

    # A number that is present but unparseable does NOT cost the student their record --
    # the parent supplied something, it just needs correcting. Storing it as NULL keeps
    # the student scannable while making them unreachable by SMS, which is the honest
    # state of affairs.
    try:
        stored = normalise(mobile)
    except InvalidMobile:
        stored = None
        notes.append(
            f"parent mobile {mobile!r} is not a PH mobile number "
            "- stored blank, this guardian will not be texted"
        )

    if guardian.casefold() == name.casefold():
        notes.append("parent name is identical to the student name")

    last, first = split_name(name)
    if not first:
        notes.append(f"name {name!r} has no comma - treated as a surname only")

    if note := surname_note(last, guardian):
        notes.append(note)

    return Candidate(
        lrn=lrn,
        first=first,
        last=last,
        section_name=section_name,
        grade_level=grade_level,
        guardian_name=guardian,
        guardian_mobile=stored,
        notes=tuple(notes),
    )


def _first_four(row: tuple) -> tuple:
    """LRN, name, parent, mobile -- padded, so a short row is not an IndexError."""
    padded = tuple(row) + (None,) * 4
    return padded[:4]


def parse_workbook(path: str | Path) -> tuple[list[Candidate], list[Rejected]]:
    """Read every worksheet. Returns (importable students, rejected rows).

    Import is deferred so that `trackify.core` stays importable on a machine without
    openpyxl -- the kiosk itself never needs it.
    """
    from openpyxl import load_workbook

    book = load_workbook(Path(path), data_only=True, read_only=True)
    students: list[Candidate] = []
    rejected: list[Rejected] = []

    for sheet in book.worksheets:
        grade_level, section_name = parse_section(sheet.title)
        for row in sheet.iter_rows(values_only=True):
            parsed = parse_row(row, grade_level, section_name)
            if isinstance(parsed, Candidate):
                students.append(parsed)
            elif isinstance(parsed, Rejected):
                rejected.append(parsed)

    book.close()
    return students, rejected
