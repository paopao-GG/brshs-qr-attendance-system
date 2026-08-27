"""DepEd School Form 2 export.

Two things are under test and they are not the same thing.

The first is GEOMETRY: this file has to land its captions on the same cells as the
school's own LIS workbook, because the form is printed, signed and submitted. A cell
that drifts one column left still opens fine and is still wrong.

The second is ARITHMETIC, and it is where SF2 departs from the register in xlsx.py --
blank means present, an excused day is not an absence, and a class-day column only
exists on evidence. Those rules are asserted directly rather than through the file's
appearance.
"""
import pytest

pytest.importorskip("openpyxl")

from openpyxl import load_workbook

from trackify.core import corrections
from trackify.core.corrections import CorrectionType
from trackify.export import sf2

WEEK = ("2026-09-01", "2026-09-02", "2026-09-03", "2026-09-04")   # Tue..Fri
YEAR, MONTH = 2026, 9


def mark(conn, student_id, day, status):
    conn.execute(
        """INSERT INTO attendance_days (student_id, date, status, created_at)
           VALUES (?, ?, ?, '2026-09-01T00:00:00+08:00')""",
        (student_id, day, status),
    )


def school_days(conn, days):
    for day in days:
        conn.execute(
            """INSERT OR IGNORE INTO school_days
               (date, is_school_day, entry_open, late_threshold,
                dismissal_time, early_departure_cutoff)
               VALUES (?, 1, '06:00', '07:15', '16:00', '15:30')""",
            (day,),
        )


@pytest.fixture
def klass(conn, make_student):
    """Two boys and one girl over four class days, one of each interesting status."""
    ben = make_student(first="Ben", last="Aquino", sex="M")
    carl = make_student(first="Carl", last="Bautista", sex="M")
    dina = make_student(first="Dina", last="Cruz", sex="F")
    school_days(conn, WEEK)

    for day in WEEK:
        mark(conn, ben, day, "present")
    mark(conn, carl, WEEK[0], "present")
    mark(conn, carl, WEEK[1], "late")
    mark(conn, carl, WEEK[2], "absent")
    mark(conn, carl, WEEK[3], "present")
    for day in WEEK:
        mark(conn, dina, day, "absent")
    return {"ben": ben, "carl": carl, "dina": dina}


@pytest.fixture
def book(conn, section, klass, config, tmp_path):
    path = sf2.export_sf2(conn, section, YEAR, MONTH, tmp_path / "sf2.xlsx",
                          config=config, school_name="Bicol Regional Science High")
    return load_workbook(path)


def row_of(sheet, name):
    """The row a learner or a block-total label sits on -- column C."""
    for row in sheet.iter_rows(min_row=sf2.FIRST_LEARNER_ROW, max_col=5):
        if row[2].value == name:
            return row[2].row
    raise AssertionError(f"{name} is not on the form")


def label_row(sheet, name):
    """The row a footer label sits on. Those live in the AM:AU panel, not column C."""
    for row in sheet.iter_rows(min_col=39, max_col=sf2.LAST_COLUMN):
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip() == name:
                return cell.row
    raise AssertionError(f"{name} is not in the summary panel")


# --- geometry ----------------------------------------------------------------
# The cells asserted here were read out of the BIFF records of the school's own
# SF2_2025_Grade-10-Year-IV-RESILIENT.xls. They are the contract.

def test_the_captions_land_where_the_deped_form_puts_them(book):
    sheet = book.active
    assert sheet["A1"].value.startswith("School Form 2 (SF2)")
    assert sheet["A3"].value.strip() == "School ID"
    assert sheet["J3"].value.strip() == "School Year"
    assert sheet["S3"].value == "Report for the Month of"
    assert sheet["A4"].value == "Name of School"
    assert sheet["S4"].value == "Grade Level"
    assert sheet["AI4"].value.strip() == "Section"
    assert sheet["A5"].value == "No."
    assert sheet["C5"].value.startswith("NAME")
    assert sheet["AM5"].value == "Total for the Month"
    assert sheet["AM7"].value == "ABSENT"
    assert sheet["AO7"].value == "PRESENT"
    assert sheet["AQ5"].value.startswith("REMARKS")


def test_the_header_carries_the_school_the_section_and_the_month(book):
    sheet = book.active
    assert sheet["F4"].value == "Bicol Regional Science High"
    assert sheet["AA3"].value == "September"
    assert sheet["AA4"].value == "Grade 7"
    assert sheet["AM4"].value == "Rizal"


def test_there_are_twenty_five_day_slots_across_forty_seven_columns():
    """Five weeks of five days, and the sheet is A..AU. Both are the form's, not ours:
    a month with more class days than slots is refused rather than truncated."""
    assert sf2.SLOTS == 25
    assert sf2.LAST_COLUMN == 47
    assert sf2.DAY_SLOTS[0] == (6, 7)        # F:G
    assert sf2.DAY_SLOTS[-1] == (37, 38)     # AK:AL


def test_the_dates_sit_above_their_weekday_letters(book):
    sheet = book.active
    assert [sheet.cell(6, first).value for first, _ in sf2.DAY_SLOTS[:4]] == [1, 2, 3, 4]
    assert [sheet.cell(7, first).value for first, _ in sf2.DAY_SLOTS[:4]] == \
        ["T", "W", "TH", "F"]


def test_unused_day_slots_are_left_empty(book):
    sheet = book.active
    for first, _ in sf2.DAY_SLOTS[len(WEEK):]:
        assert sheet.cell(6, first).value is None
        assert sheet.cell(7, first).value is None


def test_it_prints_landscape_on_one_page_wide(book):
    """The form is 47 columns. Portrait, or fitted to a page height, makes it
    unreadable -- and this is a document somebody signs on paper."""
    sheet = book.active
    assert sheet.page_setup.orientation == "landscape"
    assert sheet.page_setup.fitToWidth == 1
    assert sheet.page_setup.fitToHeight == 0


# --- the codes ---------------------------------------------------------------

def test_present_is_blank_because_sf2_marks_the_exception(book, klass):
    sheet = book.active
    line = row_of(sheet, "Aquino, Ben")
    for first, _ in sf2.DAY_SLOTS[:len(WEEK)]:
        assert sheet.cell(line, first).value is None


def test_absent_is_a_lowercase_x(book):
    sheet = book.active
    line = row_of(sheet, "Cruz, Dina")
    assert [sheet.cell(line, first).value for first, _ in sf2.DAY_SLOTS[:4]] == \
        ["x", "x", "x", "x"]


def test_late_is_shaded_and_not_marked_absent(book):
    """The form's own code for tardiness is a half-shaded cell. A late student is
    PRESENT -- shading them and also marking them 'x' would count one day twice."""
    sheet = book.active
    line = row_of(sheet, "Bautista, Carl")
    late = sheet.cell(line, sf2.DAY_SLOTS[1][0])
    assert late.value is None
    assert late.fill.fgColor.rgb == "00BFBFBF"


def test_the_monthly_totals_count_late_as_present(book):
    sheet = book.active
    line = row_of(sheet, "Bautista, Carl")
    assert sheet.cell(line, sf2.ABSENT_COL).value == 1
    assert sheet.cell(line, sf2.PRESENT_COL).value == 3


# --- excused, which is where SF2 and the register disagree -------------------

def test_an_excused_day_is_blank_and_named_in_remarks(conn, section, klass, config,
                                                      tmp_path):
    """TRACKIFY takes an excused day out of the denominator rather than counting it
    against the student (analytics-model.md 1). Marking it 'x' here would contradict
    the register the same school gets from the same database."""
    corrections.correct(conn, klass["ben"], WEEK[2], CorrectionType.EXCUSED,
                        reason="medical", actor_name="T. San Jose")
    sheet = load_workbook(
        sf2.export_sf2(conn, section, YEAR, MONTH, tmp_path / "e.xlsx", config=config)
    ).active

    line = row_of(sheet, "Aquino, Ben")
    assert sheet.cell(line, sf2.DAY_SLOTS[2][0]).value is None
    assert sheet.cell(line, sf2.ABSENT_COL).value == 0
    assert sheet.cell(line, sf2.PRESENT_COL).value == 3
    assert "Excused" in sheet.cell(line, sf2.REMARKS_COL).value
    assert "medical" in sheet.cell(line, sf2.REMARKS_COL).value


def test_a_day_with_no_record_is_blank_but_says_so(conn, section, make_student,
                                                   config, tmp_path):
    """A blank meaning 'nobody recorded this day' is not the same claim as a blank
    meaning 'this child was here', and on SF2 they look identical."""
    student = make_student(first="Eli", last="Diaz", sex="M")
    school_days(conn, WEEK)
    mark(conn, student, WEEK[0], "present")
    # WEEK[1] is a class day for the section only because another student has a row.
    other = make_student(first="Fay", last="Estrada", sex="F")
    for day in WEEK:
        mark(conn, other, day, "present")

    sheet = load_workbook(
        sf2.export_sf2(conn, section, YEAR, MONTH, tmp_path / "m.xlsx", config=config)
    ).active
    line = row_of(sheet, "Diaz, Eli")
    assert sheet.cell(line, sf2.DAY_SLOTS[1][0]).value is None
    assert "No attendance recorded" in sheet.cell(line, sf2.REMARKS_COL).value
    assert sheet.cell(line, sf2.PRESENT_COL).value == 1


# --- the male and female blocks ----------------------------------------------

def test_each_block_is_numbered_from_one(book):
    sheet = book.active
    assert sheet.cell(row_of(sheet, "Aquino, Ben"), 1).value == 1
    assert sheet.cell(row_of(sheet, "Bautista, Carl"), 1).value == 2
    assert sheet.cell(row_of(sheet, "Cruz, Dina"), 1).value == 1


def test_the_blocks_are_closed_by_their_total_rows(book):
    sheet = book.active
    labels = [sheet.cell(row, 3).value for row in range(sf2.FIRST_LEARNER_ROW, 20)]
    assert "<=== MALE | TOTAL Per Day ===>" in labels
    assert "<=== FEMALE | TOTAL Per Day ===>" in labels
    assert "Combined TOTAL Per Day" in labels


def test_the_per_day_totals_count_everyone_present_that_day(book):
    """Day 3: Ben present, Carl absent, Dina absent. Two male, one female, one present."""
    sheet = book.active
    male = row_of(sheet, "<=== MALE | TOTAL Per Day ===>")
    female = row_of(sheet, "<=== FEMALE | TOTAL Per Day ===>")
    combined = row_of(sheet, "Combined TOTAL Per Day")
    column = sf2.DAY_SLOTS[2][0]

    assert sheet.cell(male, column).value == 1
    assert sheet.cell(female, column).value == 0
    assert sheet.cell(combined, column).value == 1


def test_the_combined_row_is_the_blocks_added_up(book):
    sheet = book.active
    male = row_of(sheet, "<=== MALE | TOTAL Per Day ===>")
    female = row_of(sheet, "<=== FEMALE | TOTAL Per Day ===>")
    combined = row_of(sheet, "Combined TOTAL Per Day")
    for first, _ in sf2.DAY_SLOTS[:len(WEEK)]:
        assert sheet.cell(combined, first).value == (
            sheet.cell(male, first).value + sheet.cell(female, first).value
        )


def test_a_student_with_no_sex_recorded_is_listed_not_dropped(conn, section,
                                                              make_student, config,
                                                              tmp_path):
    """Silently omitting them would make the totals disagree with the roster, which is
    the one failure nobody would catch by looking at the form."""
    make_student(first="Gil", last="Flores", sex=None)
    school_days(conn, WEEK)
    sheet = load_workbook(
        sf2.export_sf2(conn, section, YEAR, MONTH, tmp_path / "u.xlsx", config=config)
    ).active

    assert row_of(sheet, "Flores, Gil")
    assert row_of(sheet, "<=== SEX NOT RECORDED | TOTAL Per Day ===>")


def test_the_third_block_disappears_once_everyone_is_placed(book):
    sheet = book.active
    labels = {sheet.cell(row, 3).value for row in range(sf2.FIRST_LEARNER_ROW, 20)}
    assert "<=== SEX NOT RECORDED | TOTAL Per Day ===>" not in labels


# --- which dates get a column ------------------------------------------------

def test_a_suspended_date_gets_no_column(conn, section, klass, config):
    conn.execute("UPDATE school_days SET is_school_day = 0 WHERE date = ?", (WEEK[1],))
    assert sf2.class_days(conn, section, YEAR, MONTH) == [WEEK[0], WEEK[2], WEEK[3]]


def test_a_school_day_nobody_attended_gets_no_column(conn, section, klass, config):
    """sessions.get_school_day() writes is_school_day = 1 for any date the kiosk ticks
    over, so an evening somebody opened the app becomes a school day. A column for it
    would put 'no attendance recorded' against every child and add a day to the
    divisor of the daily average."""
    school_days(conn, ["2026-09-07"])
    assert "2026-09-07" not in sf2.class_days(conn, section, YEAR, MONTH)


def test_a_weekday_the_section_missed_still_gets_a_column(conn, section, klass,
                                                          make_student):
    """The other side of that coin: the school ran and THIS section's records are what
    is missing. Hiding the day would hide the gap."""
    other = conn.execute(
        "INSERT INTO sections (name, grade_level) VALUES ('Bonifacio', 7)").lastrowid
    outsider = conn.execute(
        """INSERT INTO students (lrn, first_name, last_name, section_id, created_at)
           VALUES ('999999999999', 'Hana', 'Gomez', ?, '2026-09-01')""",
        (other,)).lastrowid
    school_days(conn, ["2026-09-07"])
    mark(conn, outsider, "2026-09-07", "present")

    assert "2026-09-07" in sf2.class_days(conn, section, YEAR, MONTH)


def test_a_weekend_needs_this_sections_own_attendance(conn, section, klass):
    """A Saturday make-up class is real. A Saturday somebody tested the scanner is not,
    and school_days cannot tell them apart."""
    saturday = "2026-09-05"
    school_days(conn, [saturday])
    assert saturday not in sf2.class_days(conn, section, YEAR, MONTH)

    mark(conn, klass["ben"], saturday, "present")
    assert saturday in sf2.class_days(conn, section, YEAR, MONTH)


def test_more_class_days_than_slots_is_refused_not_truncated(conn, section,
                                                             make_student):
    """Twenty-six days of attendance cannot be shown on a form with twenty-five
    columns. Dropping the last one silently would understate every absence in it."""
    student = make_student(first="Kit", last="Jimenez", sex="M")
    for number in range(1, 27):
        day = f"2026-09-{number:02d}"
        school_days(conn, [day])
        mark(conn, student, day, "present")

    with pytest.raises(sf2.Sf2Error, match="25"):
        sf2.class_days(conn, section, YEAR, MONTH)


# --- the summary panel -------------------------------------------------------

def test_the_summary_counts_registered_learners_by_sex(book):
    sheet = book.active
    line = label_row(sheet, "Registered Learners as of end of month")
    assert (sheet.cell(line, 44).value, sheet.cell(line, 45).value,
            sheet.cell(line, 46).value) == (2, 1, 3)


def test_average_daily_attendance_divides_by_class_days(conn, section, klass, days=None):
    """Ben 4 present, Carl 3, Dina 0 -- seven learner-days over four class days."""
    days = sf2.class_days(conn, section, YEAR, MONTH)
    stats = sf2.summary(
        [(s, [p for p in sf2.learners(conn, section, days) if p.sex == s])
         for s in (sf2.MALE, sf2.FEMALE)], days)
    assert stats["average"]["total"] == pytest.approx(7 / 4)
    assert stats["percentage"] == pytest.approx(7 / 4 / 3 * 100)


def test_boxes_trackify_cannot_answer_are_left_empty(book):
    """Enrolment as of the first Friday, transfers and drop-outs have no column in this
    database. A plausible-looking number in those boxes would be an invention on a
    document a principal signs."""
    sheet = book.active
    for label in ("* Enrolment as of (1st Friday of June)", "Dropped out",
                  "Transferred in", "Transferred out",
                  "Percentage of Enrolment as of end of month"):
        line = label_row(sheet, label)
        assert [sheet.cell(line, column).value for column in (44, 45, 46)] == \
            [None, None, None]


def test_the_number_of_days_of_classes_is_stated(book):
    sheet = book.active
    banner = [cell.value for row in sheet.iter_rows(min_col=42, max_col=42)
              for cell in row if isinstance(cell.value, str)
              and cell.value.startswith("No. of Days of Classes")]
    assert banner == ["No. of Days of Classes: 4"]


# --- five consecutive days ---------------------------------------------------

def test_a_run_of_absences_is_counted_over_class_days_not_calendar_days(conn, section,
                                                                        make_student):
    """Friday and the following Monday are consecutive in school terms. Counting the
    weekend as a break would mean a child who missed a full fortnight never triggers
    the home visitation rule in guideline 5."""
    student = make_student(first="Ivy", last="Herrera", sex="F")
    week = ["2026-09-01", "2026-09-02", "2026-09-03", "2026-09-04", "2026-09-07"]
    school_days(conn, week)
    for day in week:
        mark(conn, student, day, "absent")

    days = sf2.class_days(conn, section, YEAR, MONTH)
    learner = sf2.learners(conn, section, days)[0]
    assert sf2.consecutive_absences(learner, days) == 5


def test_one_day_back_breaks_the_run(conn, section, make_student):
    student = make_student(first="Jon", last="Ibarra", sex="M")
    week = ["2026-09-01", "2026-09-02", "2026-09-03", "2026-09-04", "2026-09-07"]
    school_days(conn, week)
    for day in week:
        mark(conn, student, day, "absent")
    conn.execute("UPDATE attendance_days SET status = 'present' WHERE date = ?",
                 ("2026-09-03",))

    days = sf2.class_days(conn, section, YEAR, MONTH)
    learner = sf2.learners(conn, section, days)[0]
    assert sf2.consecutive_absences(learner, days) == 2


# --- signatures --------------------------------------------------------------

def test_a_seeded_role_placeholder_is_not_printed_as_a_signature(conn, section,
                                                                 adviser, klass,
                                                                 config, tmp_path):
    """seed_demo seeds 'Class Adviser' so sections.adviser_id has a row to point at.
    Printed over a signature line it reads as somebody's name."""
    conn.execute("UPDATE users SET full_name = 'Class Adviser' WHERE id = ?", (adviser,))
    conn.execute("UPDATE sections SET adviser_id = ? WHERE id = ?", (adviser, section))
    sheet = load_workbook(
        sf2.export_sf2(conn, section, YEAR, MONTH, tmp_path / "s.xlsx", config=config)
    ).active

    line = label_row(sheet, "(Signature of Adviser over Printed Name)")
    assert sheet.cell(line - 1, 40).value is None


def test_a_real_adviser_is_printed(conn, section, adviser, klass, config, tmp_path):
    conn.execute("UPDATE sections SET adviser_id = ? WHERE id = ?", (adviser, section))
    sheet = load_workbook(
        sf2.export_sf2(conn, section, YEAR, MONTH, tmp_path / "s.xlsx", config=config)
    ).active

    line = label_row(sheet, "(Signature of Adviser over Printed Name)")
    assert sheet.cell(line - 1, 40).value == "Tricia San Jose"


def test_the_filename_names_the_form_the_section_and_the_month():
    assert sf2.default_filename("7-Rizal", 2026, 9) == "SF2-7-Rizal-2026-09.xlsx"
