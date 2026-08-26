"""XLSX register export.

Round-tripped through openpyxl rather than merely checked for existence: a workbook
that saves without error can still be empty, mis-shaped, or missing the corrections it
was supposed to show.
"""
import pytest

pytest.importorskip("openpyxl")

from openpyxl import load_workbook

from trackify.core import corrections
from trackify.core.attendance import record_scan
from trackify.core.corrections import CorrectionType
from trackify.export import xlsx

from .conftest import at

DAY = "2026-09-01"


@pytest.fixture
def book(conn, section, make_student, config, tmp_path):
    """One present student, one excused, exported."""
    juan = make_student(first="Juan", last="Dela Cruz")
    ana = make_student(first="Ana", last="Reyes")
    record_scan(conn, juan, at(7, 0), config)
    record_scan(conn, ana, at(7, 5), config)
    corrections.correct(conn, ana, DAY, CorrectionType.EXCUSED,
                        reason="medical", actor_name="T. San Jose")

    path = xlsx.export_register(conn, section, 2026, 9, tmp_path / "out.xlsx",
                                school_name="Bicol Regional Science High School")
    return load_workbook(path)


def _row_for(sheet, name):
    for row in sheet.iter_rows(min_row=6):
        if row[0].value == name:
            return row
    raise AssertionError(f"{name} not in the sheet")


def test_the_file_opens_and_names_the_section(book):
    sheet = book.active
    assert "Bicol Regional" in sheet["A1"].value
    assert "7-Rizal" in sheet["A2"].value
    assert "September 2026" in sheet["A2"].value


def test_every_day_of_the_month_is_a_column(book):
    sheet = book.active
    header = [c.value for c in sheet[5]]
    assert header[1:31] == list(range(1, 31))       # September has 30 days


def test_students_appear_with_their_letters(book):
    sheet = book.active
    juan = _row_for(sheet, "Dela Cruz, Juan")
    ana = _row_for(sheet, "Reyes, Ana")

    assert juan[1].value == "P"
    assert ana[1].value == "E"


def test_a_corrected_cell_is_visibly_marked(book):
    """The whole point of the audit trail is that a reader can tell which values came
    from a scan and which a person set afterwards."""
    sheet = book.active
    juan = _row_for(sheet, "Dela Cruz, Juan")
    ana = _row_for(sheet, "Reyes, Ana")

    assert ana[1].fill.fgColor.rgb.endswith("FFF3CD")
    assert not juan[1].fill.fgColor.rgb.endswith("FFF3CD")


def test_the_totals_match_the_register(conn, section, book):
    _, rows = corrections.register(conn, section, 2026, 9)
    expected = {r.name: (r.present, r.late, r.absent, r.excused) for r in rows}

    sheet = book.active
    for name, (present, late, absent, excused) in expected.items():
        row = _row_for(sheet, name)
        assert (row[31].value, row[32].value, row[33].value, row[34].value) == (
            present, late, absent, excused
        )


def test_an_excused_student_scores_full_marks_not_a_penalty(conn, section, book):
    """Excused leaves the denominator: Ana's only recorded day was excused, so her
    rate is undefined -- not 0%, which would read as never attending."""
    sheet = book.active
    ana = _row_for(sheet, "Reyes, Ana")
    juan = _row_for(sheet, "Dela Cruz, Juan")

    assert ana[35].value == "-"
    assert juan[35].value == 1.0


def test_the_legend_explains_the_excused_rule(book):
    sheet = book.active
    text = " ".join(
        str(c.value) for row in sheet.iter_rows() for c in row if c.value
    )
    assert "P present" in text
    assert "denominator" in text


def test_a_section_name_with_a_slash_does_not_break_the_sheet(
    conn, make_student, config, tmp_path
):
    """Section names are typed by teachers and will eventually contain a character
    Excel refuses in a sheet name."""
    section = conn.execute(
        "INSERT INTO sections (name, grade_level) VALUES ('A/B', 7)"
    ).lastrowid
    conn.execute(
        """INSERT INTO students (lrn, first_name, last_name, section_id,
               guardian_mobile, consent_on_file, created_at)
           VALUES ('777', 'Test', 'Student', ?, '639171234567', 1, '2026-01-01')""",
        (section,),
    )

    path = xlsx.export_register(conn, section, 2026, 9, tmp_path / "slash.xlsx")
    sheet = load_workbook(path).active
    assert "/" not in sheet.title


def test_the_default_filename_is_safe(tmp_path):
    assert xlsx.default_filename("7-Rizal", 2026, 9) == "attendance-7-Rizal-2026-09.xlsx"
    assert "/" not in xlsx.default_filename("A/B", 2026, 9)
