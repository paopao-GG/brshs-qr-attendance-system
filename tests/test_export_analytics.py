"""The analytics workbook.

The behaviour that matters most here is the empty case, because that is what the school
gets on day one: every sheet must exist and say what is missing. A missing sheet reads as
a bug and a zero reads as a finding.
"""
import pytest
from openpyxl import load_workbook

from trackify.analytics import ahp
from trackify.core.db import utcnow
from trackify.export.analytics import default_filename, export_analytics

SHEETS = ("Summary", "Trend", "Risk", "AHP", "Screening", "Model")


def record(conn, student_id, date, status, flags=""):
    conn.execute(
        """INSERT INTO attendance_days (student_id, date, status, flags, created_at)
           VALUES (?, ?, ?, ?, ?)""",
        (student_id, date, status, flags, utcnow()),
    )


def day(index: int) -> str:
    from datetime import date as Date
    from datetime import timedelta
    return (Date(2026, 9, 1) + timedelta(days=index - 1)).isoformat()


def text_of(sheet) -> str:
    return "\n".join(
        str(value) for row in sheet.iter_rows(values_only=True)
        for value in row if value is not None
    )


@pytest.fixture
def book(conn, config, tmp_path):
    def _write(**kwargs):
        path = export_analytics(conn, config, tmp_path / "analytics.xlsx", **kwargs)
        return load_workbook(path)
    return _write


# --- the empty case ---------------------------------------------------------

def test_every_sheet_exists_with_no_data(book):
    """Day one, before anyone has scanned. An absent sheet looks like a crash."""
    assert book().sheetnames == list(SHEETS)


def test_a_blocked_sheet_says_what_is_needed_not_zero(conn, book):
    sheet = book()["Trend"]
    text = text_of(sheet)

    assert "at least 3 school days" in text
    assert "Slope" not in text, "no statistic should be shown at all"
    numbers = [value for row in sheet.iter_rows(values_only=True) for value in row
               if isinstance(value, (int, float))]
    assert numbers == [], "a numeric cell here would read as a finding"


def test_the_model_sheet_explains_why_it_was_not_fitted(conn, book):
    text = text_of(book()["Model"])
    assert "No attendance has been recorded" in text
    assert "observed absence rate" in text
    assert "predicting the next day" in text


def test_the_summary_lists_the_caveats(conn, book):
    text = text_of(book()["Summary"])

    assert "PLACEHOLDER" in text
    assert "recommends review" in text.lower()
    assert "never imposes a sanction" in text.lower()


# --- with data --------------------------------------------------------------

@pytest.fixture
def cohort(conn, section, make_student):
    students = [make_student(first=f"S{n}", last=f"T{n}") for n in range(1, 6)]
    for index in range(1, 8):
        for position, student in enumerate(students):
            status = "absent" if position < index % 3 else "present"
            record(conn, student, day(index), status,
                   flags="early_departure" if position == 0 else "")
    return students


def test_the_trend_sheet_reports_every_required_statistic(conn, cohort, book):
    """Section 2: slope, intercept, R squared, p, the 95% CI, and Durbin-Watson."""
    text = text_of(book()["Trend"])

    for label in ("Slope (b)", "Intercept (a)", "R squared", "p-value for slope",
                  "95% CI lower", "95% CI upper", "Durbin-Watson"):
        assert label in text


def test_the_trend_sheet_warns_against_a_cumulative_series(conn, cohort, book):
    assert "never a running total" in text_of(book()["Trend"])


def test_the_risk_sheet_lists_students_with_their_band(conn, cohort, book):
    sheet = book()["Risk"]
    text = text_of(sheet)

    assert "Composite" in text and "Band" in text
    assert "T1, S1" in text
    assert any(band in text for band in ("Low", "Monitor", "Elevated", "High"))


def test_the_risk_sheet_labels_an_observed_rate_as_not_a_prediction(conn, cohort, book):
    """The fallback must never be mistaken for a model output."""
    text = text_of(book()["Risk"])
    assert "observed rate (model not fitted)" in text


def _book_with(conn, config, tmp_path, **risk_overrides):
    """A workbook built against a modified risk config."""
    import dataclasses

    cfg = dataclasses.replace(
        config, risk=dataclasses.replace(config.risk, **risk_overrides))
    return load_workbook(
        export_analytics(conn, cfg, tmp_path / "override.xlsx"))


def test_unattributed_band_cutoffs_are_flagged_as_placeholders(conn, cohort, config,
                                                               tmp_path):
    """A boundary decides whether a real student is referred. Nobody having set it is
    the thing that has to be said out loud."""
    book = _book_with(conn, config, tmp_path, bands_set_by="", bands_set_on="")
    assert "PLACEHOLDER" in text_of(book["Risk"]).upper()


def test_attributed_band_cutoffs_name_who_set_them(conn, cohort, config, tmp_path):
    """Once a school has set them they are not placeholders, and saying so would be
    false. The warning is replaced by the provenance, never simply dropped."""
    book = _book_with(conn, config, tmp_path,
                      bands_set_by="Guidance counsellor, BRSHS", bands_set_on="2026-08-27")
    text = text_of(book["Risk"])

    assert "Guidance counsellor, BRSHS" in text
    assert "2026-08-27" in text
    assert "PLACEHOLDER" not in text.upper()


# --- incidents on the Risk sheet ---------------------------------------------

def _incident(conn, student_id, category="bladed", severity=4,
              description="folding penknife"):
    """One confirmed prohibited-item incident, through the real table shape."""
    scan_id = conn.execute(
        """INSERT INTO scan_events (student_id, scanned_at, date, direction, method)
           VALUES (?, '2026-09-02T07:00:00', '2026-09-02', 'in', 'scan')""",
        (student_id,)).lastrowid
    event = conn.execute(
        """INSERT INTO screening_events (scan_event_id, occurred_at, metal_detected,
               outcome) VALUES (?, ?, 1, 'prohibited')""", (scan_id, utcnow())).lastrowid
    conn.execute(
        """INSERT INTO incidents (student_id, screening_event_id, occurred_at, category,
               item_description, severity)
           VALUES (?, ?, '2026-09-02T07:05:00', ?, ?, ?)""",
        (student_id, event, category, description, severity))


def test_the_risk_sheet_names_the_kind_of_item(conn, cohort, book):
    """"What kind" is the category. The free-text description is not carried."""
    _incident(conn, cohort[0], category="bladed", description="folding penknife")
    text = text_of(book()["Risk"])

    assert "bladed" in text
    assert "Max severity" in text
    assert "folding penknife" not in text,         "RA 10173: the description is the most sensitive field and adds nothing here"


def test_the_risk_sheet_says_why_a_band_was_floored(conn, cohort, book):
    _incident(conn, cohort[0], severity=4)
    text = text_of(book()["Risk"])

    assert "incident floor (severity 4)" in text
    assert "Band source" in text


def test_no_workbook_sheet_carries_an_item_description(conn, cohort, book):
    """The one string that must not appear anywhere in a file that gets emailed."""
    _incident(conn, cohort[0], description="box cutter")
    workbook = book()

    for name in workbook.sheetnames:
        assert "box cutter" not in text_of(workbook[name]), f"leaked on {name}"


# --- the AHP sheet ----------------------------------------------------------

def test_the_ahp_sheet_shows_the_consistency_check(conn, book):
    text = text_of(book()["AHP"])

    for label in ("lambda max", "Consistency Index", "Random Index", "Consistency Ratio"):
        assert label in text
    assert "0.0559" in text, "the consistency ratio itself, not just its label"


def test_unelicited_weights_are_marked_in_the_workbook(conn, book):
    text = text_of(book()["AHP"])
    assert "PLACEHOLDER WEIGHTS" in text
    assert "must not be reported as a finding" in text


def test_elicited_weights_name_their_source(conn, book):
    ahp.save(conn, ahp.DOCUMENTED_MATRIX, elicited_from="Guidance counsellor")
    text = text_of(book()["AHP"])

    assert "Guidance counsellor" in text
    assert "PLACEHOLDER WEIGHTS" not in text


# --- the screening sheet ----------------------------------------------------

def test_the_screening_sheet_names_no_student(conn, section, make_student, book):
    """RA 10173. The aggregate is counts; per-incident detail stays restricted."""
    a = make_student(first="Juan", last="Dela Cruz")
    scan_id = conn.execute(
        """INSERT INTO scan_events (student_id, scanned_at, date, direction, method)
           VALUES (?, '2026-09-01T07:00:00', '2026-09-01', 'in', 'scan')""",
        (a,)).lastrowid
    event = conn.execute(
        """INSERT INTO screening_events (scan_event_id, occurred_at, metal_detected,
               outcome) VALUES (?, ?, 1, 'prohibited')""", (scan_id, utcnow())).lastrowid
    conn.execute(
        """INSERT INTO incidents (student_id, screening_event_id, occurred_at, category,
               item_description, severity)
           VALUES (?, ?, '2026-09-01T07:05:00', 'bladed', 'penknife', 3)""",
        (a, event))

    text = text_of(book()["Screening"])
    assert "Juan" not in text
    assert "Dela Cruz" not in text
    assert "penknife" not in text
    assert "Coverage" in text


def test_the_screening_sheet_says_incidents_are_not_weighted(conn, book):
    """Both halves, because either alone is misleading.

    "Not scored" was the old wording and is now false: an incident does not enter the
    composite, but it does set a floor on the band.
    """
    text = text_of(book()["Screening"])
    assert "NOT a weighted term in the composite" in text
    assert "sets a MINIMUM band" in text
    assert "0.2212" in text, "the arithmetic, so nobody re-proposes the weighted version"


def test_the_screening_sheet_keeps_the_detail_on_the_risk_sheet(conn, book):
    assert "per-student detail is on the Risk sheet" in text_of(book()["Screening"])


# --- scope and naming -------------------------------------------------------

def test_a_section_scope_is_named_in_the_summary(conn, section, cohort, book):
    text = text_of(book(section_id=section)["Summary"])
    assert "7-Rizal" in text


def test_persist_is_off_unless_asked(conn, cohort, book):
    book()
    assert conn.execute("SELECT COUNT(*) FROM risk_scores").fetchone()[0] == 0

    book(persist=True)
    assert conn.execute("SELECT COUNT(*) FROM risk_scores").fetchone()[0] > 0


def test_the_default_filename_is_safe_for_a_filesystem():
    name = default_filename("11-Initiative/A")
    assert "/" not in name
    assert name.endswith(".xlsx")
